"""Service pembuatan laporan PDF dan Excel."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas


class ReportService:
    """Menyediakan utilitas untuk mengekspor laporan stok barang."""

    def generate_pdf(self, items: Iterable[Mapping[str, object]], destination: Path) -> None:
        pdf = canvas.Canvas(destination.as_posix(), pagesize=A4)
        width, height = A4
        pdf.setTitle("Laporan Stok Barang")

        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(2 * cm, height - 2 * cm, "Laporan Stok Barang")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(2 * cm, height - 2.7 * cm, f"Tanggal Cetak: {datetime.now():%d-%m-%Y %H:%M}")

        table_top = height - 3.5 * cm
        col_widths = [2.5 * cm, 4 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm]
        headers = ["Kode", "Nama", "Stok", "Harga Beli", "Harga Jual", "Pemasok"]

        pdf.setFillColor(colors.lightgrey)
        pdf.rect(2 * cm, table_top, sum(col_widths), 0.7 * cm, fill=1, stroke=0)
        pdf.setFillColor(colors.black)

        x = 2 * cm
        for idx, header in enumerate(headers):
            pdf.drawString(x + 0.2 * cm, table_top + 0.2 * cm, header)
            x += col_widths[idx]

        y = table_top - 0.5 * cm
        for item in items:
            if y < 3 * cm:
                pdf.showPage()
                y = height - 3 * cm
            pdf.drawString(2.2 * cm, y, str(item["item_code"]))
            pdf.drawString(4.7 * cm, y, str(item["item_name"]))
            pdf.drawRightString(2 * cm + col_widths[0] + col_widths[1] + 0.8 * cm, y, str(item["stock"]))
            pdf.drawRightString(2 * cm + sum(col_widths[:4]) - 0.2 * cm, y, f"{item['purchase_price']:.0f}")
            pdf.drawRightString(2 * cm + sum(col_widths[:5]) - 0.2 * cm, y, f"{item['selling_price']:.0f}")
            pdf.drawString(2 * cm + sum(col_widths[:5]) + 0.2 * cm, y, str(item.get("supplier", "-")))
            y -= 0.5 * cm

        pdf.showPage()
        pdf.save()

    def generate_complete_pdf(self, items: Iterable[Mapping[str, object]], transactions: Iterable[Mapping[str, object]], destination: Path) -> None:
        pdf = canvas.Canvas(destination.as_posix(), pagesize=A4)
        width, height = A4
        pdf.setTitle("Laporan Lengkap Inventori")

        # Halaman 1: Laporan Stok Barang
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(2 * cm, height - 2 * cm, "Laporan Stok Barang")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(2 * cm, height - 2.7 * cm, f"Tanggal Cetak: {datetime.now():%d-%m-%Y %H:%M}")

        table_top = height - 3.5 * cm
        col_widths = [2.5 * cm, 4 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm]
        headers = ["Kode", "Nama", "Stok", "Harga Beli", "Harga Jual", "Pemasok"]

        pdf.setFillColor(colors.lightgrey)
        pdf.rect(2 * cm, table_top, sum(col_widths), 0.7 * cm, fill=1, stroke=0)
        pdf.setFillColor(colors.black)

        x = 2 * cm
        for idx, header in enumerate(headers):
            pdf.drawString(x + 0.2 * cm, table_top + 0.2 * cm, header)
            x += col_widths[idx]

        y = table_top - 0.5 * cm
        for item in items:
            if y < 3 * cm:
                pdf.showPage()
                y = height - 3 * cm
            pdf.drawString(2.2 * cm, y, str(item["item_code"]))
            pdf.drawString(4.7 * cm, y, str(item["item_name"]))
            pdf.drawRightString(2 * cm + col_widths[0] + col_widths[1] + 0.8 * cm, y, str(item["stock"]))
            pdf.drawRightString(2 * cm + sum(col_widths[:4]) - 0.2 * cm, y, f"{item['purchase_price']:.0f}")
            pdf.drawRightString(2 * cm + sum(col_widths[:5]) - 0.2 * cm, y, f"{item['selling_price']:.0f}")
            pdf.drawString(2 * cm + sum(col_widths[:5]) + 0.2 * cm, y, str(item.get("supplier", "-")))
            y -= 0.5 * cm

        # Halaman 2: Laporan Riwayat Transaksi
        pdf.showPage()
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(2 * cm, height - 2 * cm, "Laporan Riwayat Transaksi")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(2 * cm, height - 2.7 * cm, f"Tanggal Cetak: {datetime.now():%d-%m-%Y %H:%M}")

        table_top = height - 3.5 * cm
        col_widths = [3 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 4 * cm]
        headers = ["Tanggal", "Nama Barang", "Jenis", "Qty", "Catatan"]

        pdf.setFillColor(colors.lightgrey)
        pdf.rect(2 * cm, table_top, sum(col_widths), 0.7 * cm, fill=1, stroke=0)
        pdf.setFillColor(colors.black)

        x = 2 * cm
        for idx, header in enumerate(headers):
            pdf.drawString(x + 0.2 * cm, table_top + 0.2 * cm, header)
            x += col_widths[idx]

        y = table_top - 0.5 * cm
        for transaction in transactions:
            if y < 3 * cm:
                pdf.showPage()
                y = height - 3 * cm
            pdf.drawString(2.2 * cm, y, str(transaction["transaction_date"]))
            pdf.drawString(5.2 * cm, y, str(transaction["item_name"]))
            pdf.drawString(9.2 * cm, y, str(transaction["transaction_type"]))
            pdf.drawRightString(2 * cm + sum(col_widths[:4]) - 0.2 * cm, y, str(transaction["quantity"]))
            notes = str(transaction.get("notes", "-"))
            # Truncate notes if too long for the cell
            if len(notes) > 20:
                notes = notes[:17] + "..."
            pdf.drawString(2 * cm + sum(col_widths[:4]) + 0.2 * cm, y, notes)
            y -= 0.5 * cm

        pdf.showPage()
        pdf.save()

    def generate_excel(self, items: Iterable[Mapping[str, object]], destination: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stok Barang"
        ws.append(["Kode", "Nama", "Kategori", "Stok", "Harga Beli", "Harga Jual", "Pemasok"])

        for item in items:
            ws.append(
                [
                    item["item_code"],
                    item["item_name"],
                    item.get("category", "-"),
                    item["stock"],
                    item["purchase_price"],
                    item["selling_price"],
                    item.get("supplier", "-"),
                ]
            )

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(destination.as_posix())

    def generate_transaction_pdf(self, transactions: Iterable[Mapping[str, object]], destination: Path) -> None:
        pdf = canvas.Canvas(destination.as_posix(), pagesize=A4)
        width, height = A4
        pdf.setTitle("Laporan Riwayat Transaksi")
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(2 * cm, height - 2 * cm, "Laporan Riwayat Transaksi")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(2 * cm, height - 2.7 * cm, f"Tanggal Cetak: {datetime.now():%d-%m-%Y %H:%M}")

        table_top = height - 3.5 * cm
        col_widths = [3 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 4 * cm]
        headers = ["Tanggal", "Nama Barang", "Jenis", "Qty", "Catatan"]

        pdf.setFillColor(colors.lightgrey)
        pdf.rect(2 * cm, table_top, sum(col_widths), 0.7 * cm, fill=1, stroke=0)
        pdf.setFillColor(colors.black)

        x = 2 * cm
        for idx, header in enumerate(headers):
            pdf.drawString(x + 0.2 * cm, table_top + 0.2 * cm, header)
            x += col_widths[idx]

        y = table_top - 0.5 * cm
        for transaction in transactions:
            if y < 3 * cm:
                pdf.showPage()
                y = height - 3 * cm
            pdf.drawString(2.2 * cm, y, str(transaction["transaction_date"]))
            pdf.drawString(5.2 * cm, y, str(transaction["item_name"]))
            pdf.drawString(9.2 * cm, y, str(transaction["transaction_type"]))
            pdf.drawRightString(2 * cm + sum(col_widths[:4]) - 0.2 * cm, y, str(transaction["quantity"]))
            notes = str(transaction.get("notes", "-"))
            # Truncate notes if too long for the cell
            if len(notes) > 20:
                notes = notes[:17] + "..."
            pdf.drawString(2 * cm + sum(col_widths[:4]) + 0.2 * cm, y, notes)
            y -= 0.5 * cm

        pdf.showPage()
        pdf.save()

    def generate_transaction_excel(self, transactions: Iterable[Mapping[str, object]], destination: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Riwayat Transaksi"
        ws.append(["Tanggal", "Nama Barang", "Jenis Transaksi", "Quantity", "Catatan"])

        for transaction in transactions:
            ws.append([
                transaction["transaction_date"],
                transaction["item_name"],
                transaction["transaction_type"],
                transaction["quantity"],
                transaction.get("notes", "-"),
            ])

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)  # Max width 30

        wb.save(destination.as_posix())

    def generate_complete_excel(self, items: Iterable[Mapping[str, object]], transactions: Iterable[Mapping[str, object]], destination: Path) -> None:
        wb = Workbook()

        # Sheet 1: Stok Barang
        ws1 = wb.active
        ws1.title = "Stok Barang"
        ws1.append(["Kode", "Nama", "Kategori", "Stok", "Harga Beli", "Harga Jual", "Pemasok"])

        for item in items:
            ws1.append([
                item["item_code"],
                item["item_name"],
                item.get("category", "-"),
                item["stock"],
                item["purchase_price"],
                item["selling_price"],
                item.get("supplier", "-"),
            ])

        for column_cells in ws1.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws1.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)

        # Sheet 2: Riwayat Transaksi
        ws2 = wb.create_sheet("Riwayat Transaksi")
        ws2.append(["Tanggal", "Nama Barang", "Jenis Transaksi", "Quantity", "Catatan"])

        for transaction in transactions:
            ws2.append([
                transaction["transaction_date"],
                transaction["item_name"],
                transaction["transaction_type"],
                transaction["quantity"],
                transaction.get("notes", "-"),
            ])

        for column_cells in ws2.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws2.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)

        wb.save(destination.as_posix())

